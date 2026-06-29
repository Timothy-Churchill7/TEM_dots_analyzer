#!/usr/bin/env python3
"""
TEM Quantum Dot Analyzer — Web Interface

Run:  python3 tem_webapp.py
Then open http://localhost:5000 in your browser.
"""

import io
import os
import sys
import uuid
from pathlib import Path

import numpy as np
from PIL import Image
from flask import (Flask, redirect, render_template_string,
                   request, send_file, session, url_for)
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ── Import analysis functions from the sibling script ───────────────────────
sys.path.insert(0, str(Path(__file__).parent))
from tem_dot_analyzer import (
    detect_blobs, detect_info_bar_row, detect_scale_bar_pixels,
    process_blobs, save_annotated_image,
)

# ── App setup ─────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "tem-dot-analyzer-dev-only")

WORK_DIR = Path(__file__).parent / ".tem_tmp"
WORK_DIR.mkdir(exist_ok=True)

# Server-side store: {sid -> {accumulated, file_count, current_records, ...}}
STORE: dict = {}


def get_store() -> dict:
    sid = session.get("sid")
    if not sid or sid not in STORE:
        sid = uuid.uuid4().hex
        session["sid"] = sid
        STORE[sid] = {
            "accumulated": [],   # confirmed dots across all files
            "file_count": 0,
            "current_records": [],
            "current_image_file": None,
            "current_tif_name": "",
        }
    return STORE[sid]


# ── Core analysis ─────────────────────────────────────────────────────────────

def run_analysis(tif_path: Path, scale_nm: float | None = None):
    """
    Load TIF, run dot detection and validation, save annotated PNG.
    Returns (records, annotated_png_path, nm_per_pixel or None).
    """
    img = Image.open(tif_path)
    arr = np.array(img)
    if arr.ndim == 3:
        arr = np.mean(arr[..., :3], axis=2).astype(np.uint8)

    info_bar_row = detect_info_bar_row(arr)
    nm_per_pixel = None

    if info_bar_row is not None:
        bar_px = detect_scale_bar_pixels(arr, info_bar_row)
        if bar_px:
            if scale_nm is None:
                scale_nm = 40.0 if bar_px / arr.shape[1] > 0.3 else 100.0
            nm_per_pixel = scale_nm / bar_px

    analysis = arr.copy()
    if info_bar_row is not None:
        analysis[info_bar_row:, :] = int(arr[:info_bar_row].mean())

    if nm_per_pixel:
        px_per_nm = 1.0 / nm_per_pixel
        min_r_px = (2.0 * px_per_nm) / 2.0
        max_r_px = (15.0 * px_per_nm) / 2.0
    else:
        min_r_px, max_r_px = 5.0, 150.0

    min_sigma = min_r_px / np.sqrt(2)
    max_sigma = max_r_px / np.sqrt(2)

    raw_blobs = detect_blobs(analysis, min_sigma, max_sigma,
                              threshold=0.12, overlap=0.5, denoise=1.5)
    accepted, rejected, records = process_blobs(
        raw_blobs, analysis, info_bar_row,
        nm_per_pixel=nm_per_pixel,
        min_circularity=0.65,
        max_aspect_ratio=2.0,
    )

    ann_filename = uuid.uuid4().hex + ".png"
    ann_path = WORK_DIR / ann_filename
    save_annotated_image(arr, accepted, rejected, ann_path, info_bar_row)

    img_w, img_h = int(arr.shape[1]), int(arr.shape[0])
    return records, ann_filename, nm_per_pixel, img_w, img_h


# ── Excel generation ──────────────────────────────────────────────────────────

def build_excel(accumulated: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quantum Dots"

    # Styles
    hdr_fill = PatternFill("solid", fgColor="1A2744")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EEF2F7")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        "Session #", "Source File", "File Dot #", "Confidence",
        "Length (nm)", "Length (px)", "X Position (px)", "Y Position (px)",
    ]
    col_widths = [11, 22, 12, 13, 14, 13, 17, 17]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    session_num = 0
    for row_idx, rec in enumerate(accumulated, start=2):
        session_num += 1
        values = [
            session_num,
            rec.get("source_file", ""),
            rec.get("dot_number", ""),
            rec.get("confidence", ""),
            rec.get("length_nm", ""),
            rec.get("length_px", ""),
            rec.get("centroid_x_px", ""),
            rec.get("centroid_y_px", ""),
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center
            cell.border = border
            if fill:
                cell.fill = fill

    # Summary block (2 blank rows below data)
    if accumulated:
        sizes = [r["length_nm"] for r in accumulated if r.get("length_nm") is not None]
        summary_start = len(accumulated) + 4
        stat_fill = PatternFill("solid", fgColor="FFF8E1")
        stat_hdr = Font(bold=True, size=10)
        for label, value in [
            ("Summary", ""),
            ("Total dots", len(accumulated)),
            ("Mean length (nm)", round(float(np.mean(sizes)), 3) if sizes else ""),
            ("Std dev (nm)",     round(float(np.std(sizes)),  3) if sizes else ""),
            ("Min length (nm)",  round(float(np.min(sizes)),  3) if sizes else ""),
            ("Max length (nm)",  round(float(np.max(sizes)),  3) if sizes else ""),
        ]:
            ws.cell(row=summary_start, column=1, value=label).font = stat_hdr
            ws.cell(row=summary_start, column=2, value=value)
            ws.cell(row=summary_start, column=1).fill = stat_fill
            ws.cell(row=summary_start, column=2).fill = stat_fill
            summary_start += 1

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    store = get_store()
    return render_template_string(INDEX_HTML,
        dot_count=len(store["accumulated"]),
        file_count=store["file_count"])


@app.route("/analyze", methods=["POST"])
def analyze():
    store = get_store()

    f = request.files.get("file")
    if not f or not f.filename:
        return redirect(url_for("index"))

    # Save uploaded TIF to work dir
    tif_name = Path(f.filename).name
    tif_path = WORK_DIR / (uuid.uuid4().hex + "_" + tif_name)
    f.save(tif_path)

    try:
        records, ann_filename, nm_per_pixel, img_w, img_h = run_analysis(tif_path)
    except Exception as e:
        return render_template_string(ERROR_HTML, message=str(e))
    finally:
        tif_path.unlink(missing_ok=True)

    if not records:
        return render_template_string(ERROR_HTML,
            message="No dots were detected in that image. "
                    "Try a different file or check the image format.")

    store["current_records"] = records
    store["current_image_file"] = ann_filename
    store["current_tif_name"] = tif_name

    return render_template_string(REVIEW_HTML,
        tif_name=tif_name,
        records=records,
        image_file=ann_filename,
        img_width=img_w,
        img_height=img_h,
        dot_count=len(store["accumulated"]),
        file_count=store["file_count"],
        nm_calibrated=(nm_per_pixel is not None))


@app.route("/confirm", methods=["POST"])
def confirm():
    store = get_store()

    valid_nums = set()
    for v in request.form.getlist("valid_dots"):
        try:
            valid_nums.add(int(v))
        except ValueError:
            pass

    tif_name = store["current_tif_name"]
    for rec in store["current_records"]:
        if rec["dot_number"] in valid_nums:
            store["accumulated"].append({**rec, "source_file": tif_name})

    store["file_count"] += 1
    confirmed = sum(1 for r in store["accumulated"]
                    if r.get("source_file") == tif_name)

    return render_template_string(DONE_HTML,
        tif_name=tif_name,
        confirmed=confirmed,
        total_dots=len(store["accumulated"]),
        total_files=store["file_count"])


@app.route("/download")
def download():
    store = get_store()
    if not store["accumulated"]:
        return redirect(url_for("index"))
    buf = build_excel(store["accumulated"])
    return send_file(buf, as_attachment=True,
                     download_name="quantum_dots.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/image/<filename>")
def serve_image(filename):
    path = WORK_DIR / filename
    if not path.exists():
        return "Image not found", 404
    return send_file(path, mimetype="image/png")


@app.route("/reset", methods=["POST"])
def reset():
    sid = session.get("sid")
    if sid and sid in STORE:
        del STORE[sid]
    session.clear()
    return redirect(url_for("index"))


# ── HTML Templates ────────────────────────────────────────────────────────────

_BASE_CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
  background: #f0f2f5;
  color: #1a1a2e;
  min-height: 100vh;
}
header {
  background: #1a2744;
  color: white;
  padding: 0 32px;
  height: 56px;
  display: flex;
  align-items: center;
  justify-content: space-between;
  box-shadow: 0 2px 8px rgba(0,0,0,0.3);
}
header h1 { font-size: 1.15rem; letter-spacing: 0.04em; font-weight: 600; }
header .subtitle { font-size: 0.8rem; color: #90a4ae; margin-top: 2px; }
.container { max-width: 960px; margin: 0 auto; padding: 32px 20px; }
.container.wide { max-width: 1200px; }
.card {
  background: white;
  border-radius: 10px;
  box-shadow: 0 2px 12px rgba(0,0,0,0.08);
  padding: 28px 32px;
  margin-bottom: 24px;
}
.card h2 { font-size: 1.1rem; font-weight: 600; margin-bottom: 16px; color: #1a2744; }
.badge {
  display: inline-flex; align-items: center; gap: 8px;
  background: #e8f5e9; border: 1px solid #a5d6a7; border-radius: 6px;
  padding: 10px 16px; color: #2e7d32; font-size: 0.9rem; margin-bottom: 20px;
}
.badge svg { flex-shrink: 0; }
.btn {
  display: inline-flex; align-items: center; justify-content: center; gap: 8px;
  padding: 11px 24px; border-radius: 7px; font-size: 0.95rem;
  font-weight: 600; cursor: pointer; border: none; text-decoration: none;
  transition: background 0.15s, transform 0.1s; user-select: none;
}
.btn:active { transform: scale(0.98); }
.btn-primary { background: #1976d2; color: white; }
.btn-primary:hover { background: #1565c0; }
.btn-success { background: #2e7d32; color: white; }
.btn-success:hover { background: #1b5e20; }
.btn-outline {
  background: white; color: #1976d2;
  border: 2px solid #1976d2;
}
.btn-outline:hover { background: #e3f2fd; }
.btn-danger { background: #c62828; color: white; }
.btn-danger:hover { background: #b71c1c; }
.btn-lg { padding: 15px 32px; font-size: 1.05rem; border-radius: 9px; }
/* Loading overlay */
#loading-overlay {
  display: none; position: fixed; inset: 0;
  background: rgba(10, 15, 30, 0.75); z-index: 999;
  align-items: center; justify-content: center; flex-direction: column;
  color: white; gap: 16px;
}
.spinner {
  width: 48px; height: 48px; border: 4px solid rgba(255,255,255,0.2);
  border-top-color: white; border-radius: 50%;
  animation: spin 0.9s linear infinite;
}
@keyframes spin { to { transform: rotate(360deg); } }
"""

INDEX_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>TEM Dot Analyzer</title>
<style>
{{ css }}
.upload-zone {
  border: 2.5px dashed #90a4ae; border-radius: 10px;
  padding: 56px 32px; text-align: center; cursor: pointer;
  transition: border-color 0.2s, background 0.2s;
}
.upload-zone:hover, .upload-zone.drag-over {
  border-color: #1976d2; background: #e3f2fd;
}
.upload-zone input[type=file] { display: none; }
.upload-icon { font-size: 3rem; margin-bottom: 12px; }
.upload-zone h3 { font-size: 1.1rem; color: #37474f; margin-bottom: 6px; }
.upload-zone p  { font-size: 0.85rem; color: #78909c; }
.reset-form { display: inline; }
</style>
</head>
<body>
<header>
  <div>
    <h1>TEM Quantum Dot Analyzer</h1>
    <div class="subtitle">Upload · Review · Export</div>
  </div>
  {% if dot_count > 0 %}
  <form action="/reset" method="post" class="reset-form">
    <button type="submit" class="btn btn-danger" style="font-size:0.8rem;padding:7px 14px"
      onclick="return confirm('Clear all session data and start over?')">
      Clear Session
    </button>
  </form>
  {% endif %}
</header>

<div class="container">
  {% if dot_count > 0 %}
  <div class="badge">
    <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
      <polyline points="20 6 9 17 4 12"></polyline>
    </svg>
    Session active — <strong>{{ dot_count }} dot{{ 's' if dot_count != 1 else '' }}</strong>
    confirmed from <strong>{{ file_count }} image{{ 's' if file_count != 1 else '' }}</strong>
    &nbsp;·&nbsp;
    <a href="/download" class="btn btn-success" style="padding:4px 12px;font-size:0.8rem;">
      Download Excel
    </a>
  </div>
  {% endif %}

  <div class="card">
    <h2>{% if dot_count > 0 %}Add Another Image{% else %}Upload TEM Image{% endif %}</h2>
    <p style="color:#546e7a;margin-bottom:20px;font-size:0.93rem;">
      Upload a <strong>.tif</strong> or <strong>.tiff</strong> TEM scan.
      The analyzer will detect and validate quantum dots automatically.
      You'll then review and approve which dots to include.
    </p>

    <form id="upload-form" action="/analyze" method="post" enctype="multipart/form-data">
      <div class="upload-zone" id="upload-zone"
           onclick="document.getElementById('file-input').click()">
        <input type="file" id="file-input" name="file" accept=".tif,.tiff">
        <div class="upload-icon">🔬</div>
        <h3>Click to choose a TIF file</h3>
        <p>or drag and drop here</p>
        <p style="margin-top:8px;color:#b0bec5">Analysis takes ~20 seconds per image</p>
      </div>
    </form>
  </div>

  {% if dot_count > 0 %}
  <div style="text-align:center;margin-top:8px;">
    <a href="/download" class="btn btn-success btn-lg">
      ⬇ Download Excel ({{ dot_count }} dot{{ 's' if dot_count != 1 else '' }})
    </a>
  </div>
  {% endif %}
</div>

<div id="loading-overlay">
  <div class="spinner"></div>
  <div style="font-size:1.15rem;font-weight:600;">Analyzing image…</div>
  <div style="font-size:0.85rem;color:#b0bec5;">Detecting and validating dots — about 20 seconds</div>
</div>

<script>
const input = document.getElementById('file-input');
const zone  = document.getElementById('upload-zone');
const overlay = document.getElementById('loading-overlay');
const form  = document.getElementById('upload-form');

input.addEventListener('change', () => {
  if (input.files.length) {
    overlay.style.display = 'flex';
    form.submit();
  }
});

zone.addEventListener('dragover', e => { e.preventDefault(); zone.classList.add('drag-over'); });
zone.addEventListener('dragleave', () => zone.classList.remove('drag-over'));
zone.addEventListener('drop', e => {
  e.preventDefault(); zone.classList.remove('drag-over');
  const dt = e.dataTransfer;
  if (dt.files.length) {
    // Transfer dropped file to the input and submit
    const transfer = new DataTransfer();
    transfer.items.add(dt.files[0]);
    input.files = transfer.files;
    overlay.style.display = 'flex';
    form.submit();
  }
});
</script>
</body>
</html>
""".replace("{{ css }}", _BASE_CSS)


REVIEW_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Review Dots — {{ tif_name }}</title>
<style>
{{ css }}
.review-layout { display: grid; grid-template-columns: 1fr 340px; gap: 24px; align-items: start; }
@media (max-width: 768px) { .review-layout { grid-template-columns: 1fr; } }
.image-container {
  position: relative; line-height: 0;
  border-radius: 8px; overflow: hidden;
  box-shadow: 0 2px 12px rgba(0,0,0,0.15);
  background: #111; cursor: zoom-in;
}
.image-container img { width: 100%; display: block; }
/* SVG dot circles */
.dot-circle { transition: stroke-dasharray 0.15s, opacity 0.15s; }
.dot-circle.deselected { stroke-dasharray: 7 5; opacity: 0.4; }
/* Dot grid */
.dot-grid {
  display: grid; grid-template-columns: repeat(auto-fill, minmax(72px, 1fr));
  gap: 8px; margin-top: 12px; max-height: 440px; overflow-y: auto;
  padding-right: 4px;
}
.dot-card {
  display: flex; flex-direction: column; align-items: center;
  border: 2px solid #c8e6c9; border-radius: 8px; padding: 8px 4px 6px;
  cursor: pointer; transition: all 0.15s; background: #f1f8e9;
  user-select: none;
}
.dot-card input[type=checkbox] { display: none; }
.dot-card .num {
  font-size: 1.2rem; font-weight: 700; color: #2e7d32; line-height: 1;
}
.dot-card .size { font-size: 0.68rem; color: #558b2f; margin-top: 2px; }
/* Confidence bar */
.conf-bar {
  width: 80%; height: 4px; background: rgba(0,0,0,0.08);
  border-radius: 2px; margin: 4px 0 1px; overflow: hidden;
}
.conf-fill { height: 100%; border-radius: 2px; background: #43a047; }
.conf-pct { font-size: 0.6rem; color: #78909c; }
.dot-card:not(.checked) {
  border-color: #ffcdd2; background: #fff3f3;
}
.dot-card:not(.checked) .num { color: #c62828; }
.dot-card:not(.checked) .size { color: #e57373; }
.dot-card:not(.checked) .conf-fill { background: #ef9a9a; }
.dot-card.checked { border-color: #43a047; background: #e8f5e9; }
.instructions {
  font-size: 0.85rem; color: #546e7a; line-height: 1.5;
  margin-bottom: 14px;
}
.tally {
  font-size: 0.9rem; font-weight: 600; color: #1a2744;
  margin-bottom: 12px;
}
.select-btns { display: flex; gap: 8px; margin-bottom: 12px; }
.select-btns button {
  font-size: 0.78rem; padding: 5px 10px; border-radius: 5px;
  border: 1px solid #b0bec5; background: white; cursor: pointer; color: #37474f;
}
.select-btns button:hover { background: #eceff1; }
</style>
</head>
<body>
<header>
  <div>
    <h1>TEM Quantum Dot Analyzer</h1>
    <div class="subtitle">Review detected dots — {{ tif_name }}</div>
  </div>
</header>

<div class="container wide" style="padding-top:24px;">
  {% if dot_count > 0 %}
  <div class="badge" style="margin-bottom:20px;">
    <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
      <polyline points="20 6 9 17 4 12"></polyline>
    </svg>
    Session: {{ dot_count }} dot{{ 's' if dot_count != 1 else '' }} already confirmed from {{ file_count }} image{{ 's' if file_count != 1 else '' }}
  </div>
  {% endif %}

  <form action="/confirm" method="post" id="confirm-form">

  <div class="review-layout">
    <!-- Left: annotated image with SVG circle overlay -->
    <div>
      <div class="image-container" onclick="window.open('/image/{{ image_file }}','_blank')">
        <img src="/image/{{ image_file }}" alt="TEM image" title="Click to open full size in new tab">
        <svg viewBox="0 0 {{ img_width }} {{ img_height }}"
             preserveAspectRatio="xMidYMid meet"
             style="position:absolute;top:0;left:0;width:100%;height:100%;pointer-events:none">
          {% for rec in records %}
          {% set r = rec.length_px / 2 + 5 %}
          <circle id="svgdot-{{ rec.dot_number }}"
                  class="dot-circle"
                  cx="{{ rec.centroid_x_px }}"
                  cy="{{ rec.centroid_y_px }}"
                  r="{{ r }}"
                  fill="none"
                  stroke="#22dd22"
                  stroke-width="1"
                  vector-effect="non-scaling-stroke"/>
          {% endfor %}
        </svg>
      </div>
      <p style="font-size:0.75rem;color:#90a4ae;margin-top:6px;text-align:center;">
        Click image to open full size &nbsp;·&nbsp;
        Solid circle = selected &nbsp;·&nbsp; Dashed = deselected
      </p>
    </div>

    <!-- Right: dot selection panel -->
    <div>
      <div class="card" style="padding:20px;">
        <h2 style="margin-bottom:8px;">
          Select Valid Dots
        </h2>
        <p class="instructions">
          <strong>{{ records|length }}</strong> dots detected, ranked by confidence
          (#1 = most circular &amp; isolated). The bar under each number shows
          relative confidence. Click any card to deselect it.
        </p>

        <div class="tally" id="tally">{{ records|length }} / {{ records|length }} selected</div>

        <div class="select-btns">
          <button type="button" onclick="selectAll()">Select all</button>
          <button type="button" onclick="selectNone()">Deselect all</button>
        </div>

        <div class="dot-grid" id="dot-grid">
          {% for rec in records %}
          {% set conf_pct = (rec.get('confidence', 0) * 100) | int %}
          <label class="dot-card checked" data-num="{{ rec.dot_number }}">
            <input type="checkbox" name="valid_dots"
                   value="{{ rec.dot_number }}" checked>
            <span class="num">{{ rec.dot_number }}</span>
            <div class="conf-bar">
              <div class="conf-fill" style="width:{{ conf_pct }}%"></div>
            </div>
            <span class="conf-pct">{{ conf_pct }}%</span>
            <span class="size">
              {% if rec.get('length_nm') %}{{ "%.2f"|format(rec.length_nm) }} nm{% else %}{{ "%.0f"|format(rec.length_px) }} px{% endif %}
            </span>
          </label>
          {% endfor %}
        </div>

        <div style="margin-top:16px; border-top:1px solid #e0e0e0; padding-top:16px;">
          <button type="submit" class="btn btn-primary" style="width:100%;justify-content:center;">
            Confirm Selection →
          </button>
        </div>
      </div>
    </div>
  </div>

  </form>
</div>

<script>
function updateTally() {
  const boxes = document.querySelectorAll('input[name="valid_dots"]');
  const checked = [...boxes].filter(b => b.checked).length;
  document.getElementById('tally').textContent = checked + ' / ' + boxes.length + ' selected';
}

function toggleCircle(num, checked) {
  const c = document.getElementById('svgdot-' + num);
  if (c) c.classList.toggle('deselected', !checked);
}

document.querySelectorAll('.dot-card').forEach(card => {
  card.addEventListener('click', e => {
    const num = parseInt(card.dataset.num);
    const cb = card.querySelector('input[type=checkbox]');
    cb.checked = !cb.checked;
    card.classList.toggle('checked', cb.checked);
    toggleCircle(num, cb.checked);
    updateTally();
    e.preventDefault();
  });
});

function selectAll() {
  document.querySelectorAll('.dot-card').forEach(card => {
    const num = parseInt(card.dataset.num);
    card.querySelector('input').checked = true;
    card.classList.add('checked');
    toggleCircle(num, true);
  });
  updateTally();
}
function selectNone() {
  document.querySelectorAll('.dot-card').forEach(card => {
    const num = parseInt(card.dataset.num);
    card.querySelector('input').checked = false;
    card.classList.remove('checked');
    toggleCircle(num, false);
  });
  updateTally();
}
</script>
</body>
</html>
""".replace("{{ css }}", _BASE_CSS)


DONE_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Dots Confirmed</title>
<style>
{{ css }}
.result-card { text-align: center; padding: 40px 32px; }
.check-icon { font-size: 3.5rem; margin-bottom: 16px; }
.stat-row { display: flex; justify-content: center; gap: 40px; margin: 24px 0; }
.stat { display: flex; flex-direction: column; align-items: center; }
.stat .value { font-size: 2.2rem; font-weight: 700; color: #1976d2; line-height: 1; }
.stat .label { font-size: 0.8rem; color: #78909c; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.05em; }
.action-row { display: flex; gap: 16px; justify-content: center; margin-top: 28px; flex-wrap: wrap; }
</style>
</head>
<body>
<header>
  <div>
    <h1>TEM Quantum Dot Analyzer</h1>
    <div class="subtitle">Session summary</div>
  </div>
</header>

<div class="container">
  <div class="card result-card">
    <div class="check-icon">✅</div>
    <h2 style="font-size:1.3rem;margin-bottom:6px;">
      {{ confirmed }} dot{{ 's' if confirmed != 1 else '' }} confirmed from<br>
      <span style="color:#1976d2;">{{ tif_name }}</span>
    </h2>

    <div class="stat-row">
      <div class="stat">
        <div class="value">{{ total_dots }}</div>
        <div class="label">Total Dots</div>
      </div>
      <div class="stat">
        <div class="value">{{ total_files }}</div>
        <div class="label">Image{{ 's' if total_files != 1 else '' }} Processed</div>
      </div>
    </div>

    <div class="action-row">
      <a href="/download" class="btn btn-success btn-lg">
        ⬇ Download Excel
      </a>
      <a href="/" class="btn btn-outline btn-lg">
        ＋ Add Another TIF
      </a>
    </div>
  </div>
</div>
</body>
</html>
""".replace("{{ css }}", _BASE_CSS)


ERROR_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Error</title>
<style>{{ css }}</style>
</head>
<body>
<header><h1>TEM Quantum Dot Analyzer</h1></header>
<div class="container">
  <div class="card" style="text-align:center;padding:40px;">
    <div style="font-size:3rem;margin-bottom:16px;">⚠️</div>
    <h2 style="color:#c62828;margin-bottom:12px;">Analysis Failed</h2>
    <p style="color:#546e7a;max-width:480px;margin:0 auto 24px;">{{ message }}</p>
    <a href="/" class="btn btn-primary">← Try Another File</a>
  </div>
</div>
</body>
</html>
""".replace("{{ css }}", _BASE_CSS)


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    print("\n  TEM Dot Analyzer — Web Interface")
    print(f"  Open http://localhost:{port} in your browser\n")
    app.run(debug=False, host="0.0.0.0", port=port)
